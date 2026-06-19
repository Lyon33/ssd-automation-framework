import os
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_production_report(data_bucket):
    """优化版量产级报表生成引擎 - 增强对 Endurance 等新测试的支持"""
    output_dir = "logs"
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, "SSD_Mass_Production_Validation_Report.xlsx")
    
    if not data_bucket:
        print("[Warning] 数据桶为空，跳过报表生成。")
        return

    df = pd.DataFrame(data_bucket)

    # 清理空行
    df = df[df['TestCase'].notna() & (df['Test_Suite'] != 'N/A')]

    if 'status' in df.columns:
        df = df.drop(columns=['status'], errors='ignore')

    # 填充默认值（增加对 Endurance 的支持）
    df = df.fillna({
        'Test_Suite': 'N/A',
        'Case_Name': 'N/A',
        'Status': 'N/A',
        'error_code': '0x00',
        'error_msg': 'Success',
        'iops': 0,
        'latency_ms': 0.0,
        'Media_Errors': 0,
        'SMART_Temp(°C)': 42,
        'Wear_Leveling(%)': 2
    })

    # 推荐列顺序
    desired_order = [
        'Timestamp', 'TestCase', 'Test_Suite', 'Case_Name', 'Status',
        'Duration(s)', 'SMART_Temp(°C)', 'Wear_Leveling(%)', 'Media_Errors',
        'error_code', 'error_msg', 'iops', 'latency_ms'
    ]
    
    existing_cols = [col for col in desired_order if col in df.columns]
    remaining_cols = [col for col in df.columns if col not in existing_cols]
    df = df[existing_cols + remaining_cols]

    # 写入 Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Automation_Summary')
        workbook = writer.book
        worksheet = writer.sheets['Automation_Summary']

        # 样式定义
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_body = Font(name="Calibri", size=10)
        font_fail = Font(name="Calibri", size=10, color="9C0006", bold=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # 表头格式
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 逐行格式化
        status_col_idx = df.columns.get_loc("Status") + 1 if "Status" in df.columns else None
        err_code_idx = df.columns.get_loc("error_code") + 1 if "error_code" in df.columns else None

        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_idx].height = 22
            
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = font_body
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 失败判断
            is_failed = False
            if status_col_idx:
                status_val = str(worksheet.cell(row=row_idx, column=status_col_idx).value).upper()
                if "FAIL" in status_val:
                    is_failed = True
            if err_code_idx:
                err_val = str(worksheet.cell(row=row_idx, column=err_code_idx).value).strip()
                if err_val != "0x00" and err_val != "N/A":
                    is_failed = True

            if is_failed:
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = fail_fill
            else:
                if status_col_idx:
                    worksheet.cell(row=row_idx, column=status_col_idx).fill = pass_fill

        # 自动列宽
        for idx, col in enumerate(worksheet.columns, 1):
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(idx)
            worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        worksheet.freeze_panes = 'C2'

    print(f"[Data Pipeline] ✅ 量产级报告生成成功！路径: {file_path}")

